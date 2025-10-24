#!/usr/bin/env python3
"""
Convert samples.xlsx to sample_books.csv with proper UTF-8 encoding.
This fixes encoding issues like "Â®", "â€™", etc.

Usage:
    python3 convert_excel_to_csv.py

Requirements:
    pip install openpyxl
"""

import csv
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed.")
    print("Install with: pip3 install openpyxl")
    exit(1)


def clean_text(text: str) -> str:
    """
    Clean Excel XML encoding artifacts and fix special characters.
    Handles both Excel XML escapes and UTF-8 mojibake (mis-encoded characters).
    """
    if not text or text == 'None':
        return ''

    # Fix Excel XML escape sequences
    xml_replacements = {
        'â_x0080__x0099_': "'",  # Right single quote
        'â_x0080__x0098_': "'",  # Left single quote
        'â_x0080__x009c_': '"',  # Left double quote
        'â_x0080__x009d_': '"',  # Right double quote
        'â_x0080__x0093_': '–',  # En dash
        'â_x0080__x0094_': '—',  # Em dash
        '_x0080__x0099_': "'",
        '_x0080__x0098_': "'",
        '_x0080__x009c_': '"',
        '_x0080__x009d_': '"',
        '_x0080__x0093_': '–',
        '_x0080__x0094_': '—',
        'Ã_x009f_': 'ß',  # German eszett (ß) with XML escape
        '_x009f_': 'ß',   # Alternative format
    }

    for corrupted, correct in xml_replacements.items():
        text = text.replace(corrupted, correct)

    # Fix German and other European character mojibake
    # These are UTF-8 bytes interpreted as Windows-1252/ISO-8859-1
    mojibake_replacements = {
        # German umlauts and eszett
        'Ã¤': 'ä',  # a umlaut
        'Ã¶': 'ö',  # o umlaut
        'Ã¼': 'ü',  # u umlaut
        'Ã„': 'Ä',  # A umlaut
        'Ã–': 'Ö',  # O umlaut
        'Ãœ': 'Ü',  # U umlaut
        'ÃŸ': 'ß',  # eszett (sharp s)

        # French accents
        'Ã©': 'é',  # e acute
        'Ã¨': 'è',  # e grave
        'Ãª': 'ê',  # e circumflex
        'Ã«': 'ë',  # e diaeresis
        'Ã¢': 'â',  # a circumflex
        'Ã ': 'à',  # a grave
        'Ã§': 'ç',  # c cedilla
        'Ã®': 'î',  # i circumflex
        'Ã´': 'ô',  # o circumflex

        # Spanish
        'Ã±': 'ñ',  # n tilde
        'Ñ': 'Ñ',  # N tilde
        'Ã³': 'ó',  # o acute
        'Ã­': 'í',  # i acute
        'Ãº': 'ú',  # u acute
        'Ã¡': 'á',  # a acute

        # Special symbols
        'Â®': '®',   # registered trademark
        'â„¢': '™',  # trademark
        'Â©': '©',   # copyright
        'Â°': '°',   # degree
        'Â': '',     # non-breaking space artifact (when alone)

        # Smart quotes and dashes
        'â€™': "'",  # right single quote
        'â€˜': "'",  # left single quote
        'â€œ': '"',  # left double quote
        'â€': '"',   # right double quote
        'â€"': '—',  # em dash
        'â€"': '–',  # en dash
        'â€¦': '…',  # ellipsis
    }

    for corrupted, correct in mojibake_replacements.items():
        text = text.replace(corrupted, correct)

    return text


def convert_excel_to_csv(excel_path: str, csv_path: str, sheet_name: str = None):
    """
    Convert Excel file to CSV with proper UTF-8 encoding.

    Args:
        excel_path: Path to input .xlsx file
        csv_path: Path to output .csv file
        sheet_name: Name of sheet to export (default: first sheet)
    """
    print(f"Loading Excel file: {excel_path}")

    # Load workbook
    wb = load_workbook(excel_path, data_only=True)

    # Select sheet
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    print(f"Exporting sheet: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")

    # Write to CSV with proper UTF-8 encoding
    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile, quoting=csv.QUOTE_MINIMAL)

        row_count = 0
        for row in ws.iter_rows(values_only=True):
            # Convert None to empty string and clean text artifacts
            clean_row = [clean_text(str(cell)) if cell is not None else '' for cell in row]
            writer.writerow(clean_row)
            row_count += 1

    print(f"✓ Successfully exported {row_count} rows to: {csv_path}")
    print(f"✓ File saved with UTF-8 encoding")
    print(f"✓ Cleaned Excel XML encoding artifacts")


def main():
    # Paths
    script_dir = Path(__file__).parent
    app_dir = script_dir / "MiniLibrary" / "MiniLibrary" / "App"

    excel_file = app_dir / "samples.xlsx"
    csv_file = app_dir / "sample_books.csv"

    # Check if Excel file exists
    if not excel_file.exists():
        print(f"Error: Excel file not found at {excel_file}")
        print("Please ensure samples.xlsx is in the App directory.")
        exit(1)

    # Convert
    convert_excel_to_csv(str(excel_file), str(csv_file))

    print("\n✅ Conversion complete!")
    print("The CSV file is now ready to use with proper UTF-8 encoding.")
    print("All special characters (®, ™, ö, ä, etc.) should display correctly.")


if __name__ == "__main__":
    main()
